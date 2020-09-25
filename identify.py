import cognitive_face as CF
from global_variables import personGroupId
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, cell, column_index_from_string
import time


wb = load_workbook(filename="reports.xlsx")
currentDate = time.strftime("%d_%m_%y")

def update_attendance(student_roll):
	if(os.path.exists('./reports.xlsx')):
		wb = load_workbook(filename = "reports.xlsx")
		sheet = wb['IT2016']
		if sheet.cell(row=1, column=3).value == currentDate:
			for row in range(2,sheet.max_row+1):
				if str(sheet.cell(row=row, column=1).value) == student_roll:
					ans='present'
					sheet.cell(row=row,column=3).value = ans
					wb.save(filename = "reports.xlsx")
					break 

Base_url = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0' # default endpoint when use free 
CF.BaseUrl.set(Base_url)
Key= '20c5c11712244ba4b340ac8bf949b8ef'
CF.Key.set(Key)
			
currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')   # it is the directory where the faces are cropped from photo provided by pics folder 
for filename in os.listdir(directory):  # going through each and every images present in the photo 
	if filename.endswith(".jpg"):
		imgurl = urllib.request.pathname2url(os.path.join(directory, filename)) 
		res = CF.face.detect(imgurl)   #creating the faceId for every images 
		#print(res)
		if len(res) != 1:
			print("No face detected.")
			continue
	    # res will contain all the details of the face detected like --
	    # [{'faceId': '450ae3a9-6413-40a4-b945-f94b4c6a2874', 'faceRectangle': {'top': 32, 'left': 17, 'width': 352, 'height': 352}}]
		faceIds = []
		for face in res:
			faceIds.append(face['faceId']) # faceId to faceIds list 
		#print(faceIds)
		res = CF.face.identify(faceIds, personGroupId) # you have provided the faceIds to  CF.face.identify the faces 
		print(res)
		for face in res:
			if not face['candidates']:
				print("Unknown")
			else:
				personId = face['candidates'][0]['personId']
				print(personId)
				connect = sqlite3.connect('face-database.db')
				c = connect.cursor()
				c.execute("select * from students where personId = ?",(personId,))
				row = c.fetchone()
				if row==None:
					print("Attendance Updation Failed")
					break
				student_roll = row[2]
				print(student_roll)
				update_attendance(student_roll)
				print(row[1]+" recognised")
			time.sleep(6)













